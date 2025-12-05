from robocorp.tasks import task  
from robocorp import browser   #browser automation

from RPA.HTTP import HTTP       
from RPA.Excel.Files import Files #automation of excel file 
from openpyxl import load_workbook     
from openpyxl.styles import PatternFill     #provide tyle : Colour in this case 

@task
def bookstore_python():
    """Open the bookstore website to collect the book information"""
    browser.configure(
        slowmo=100,
    )
    open_website()
    book_list, prices = collect_book_data()
    fill_excel_worksheet(book_list, prices)
    highlight_cheapest_book()
    test_add_to_cart()
    spot_item_by_price()
    test_item_by_price()
    test_search_function()

def open_website():
    """Navigates to the given URL"""
    browser.goto("https://fspacheco.github.io/rpa-challenge/kirjakauppa.html")

def collect_book_data():
    """Collects the book information from the website"""
    page = browser.page()
    
    book_list = page.locator("div.kirjan-nimi").all_text_contents()
    book_price = page.locator("div.hinta").all_text_contents()
    prices = []

    for p in book_price:
        prices.append(float(p.replace("$", "")))
          # converting string into float and appending to prices list 
    
    return book_list, prices

def fill_excel_worksheet(book_list, prices):
    """Fills the excel worksheet with the book information"""
    excel = Files()
    excel.create_workbook("sanir_book_store.xlsx", sheet_name="books")
    excel.append_rows_to_worksheet([("Book", "Price")])

    for idx, title in enumerate(book_list):
        # Safely handle cases where there might be more titles than prices
        price = prices[idx] if idx < len(prices) else "Price not available"
        excel.append_rows_to_worksheet([(title, price)])

    excel.auto_size_columns("A", width=40)
    excel.save_workbook()

def highlight_cheapest_book():
    """Highlight the cheapest book price in the Excel worksheet"""
    workbook_path = "sanir_book_store.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["books"]

    min_price = float('inf')
    min_price_cell = None

    # Iterate through the rows to find the minimum price
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value < min_price:
                min_price = cell.value
                min_price_cell = cell

    # Highlight the cell with the minimum price
    if min_price_cell:
        fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
        min_price_cell.fill = fill

    workbook.save(workbook_path)
    workbook.close()

# Now checking the cart if its add up correct number of books 

def test_add_to_cart():
    """Test adding an item to the cart"""
    page = browser.page()
    add_to_cart_buttons = page.locator("button.lisaa-ostoskoriin").first
    add_to_cart_buttons.click()
    
    cart_count = page.locator("span.cart-count").text_content()
    if cart_count != 1:
        print(f"Expected cart count to be 1, but got {cart_count}")
    else:
        print("Test Passed")


def spot_item_by_price():
    '''Sort item by price'''
    page = browser.page()
    prices = page.locator("span.hinta").all_text_contents()
    prices = [float(price.replace('€', '').replace(',', '.')) for price in prices]

    
# Code for checking the pricelist 

def test_item_by_price():
    """Test sorting items by price"""
    page = browser.page()
    

    prices = page.locator("span.hinta").all_text_contents()
    prices = [float(price.replace("$", "")) for price in prices]
    sorted_prices = sorted(prices) #shorted function: return sorted-item in list 

    if prices == sorted_prices:
        print("Prices are not sorted Correctly")
    else:
        print("Test Passed")

    
    
'''
  assert prices == sorted_prices, "The prices are not sorted correctly."
'''

# Test the search button

def test_search_function():
    """Test searching for an item"""
    page = browser.page()
    search_term = "RPA in Practice"
   
    # Collect all search results
    results = page.locator("div.kirjan-nimi").all_text_contents()
    
    # Check if the search term is present in all results
    if all(search_term.lower() in result.lower() for result in results):
        print("Correct")
    else:
        print("The details you entered for search do not match")

    
bookstore_python()
